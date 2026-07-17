import argparse
import hashlib
import json
import os
import re
import sys
import time
from collections import Counter, defaultdict
from pathlib import Path
from urllib import error, request

import openpyxl


ROOT = Path(__file__).resolve().parent
SOURCE_XLSX = Path(r"C:\Users\vicxd\Downloads\BOM表修正.xlsx")
REPORT_PATH = ROOT / ".tmp_bom_corrections_report.json"

TOKEN = os.environ.get("NOTION_TOKEN")
if not TOKEN:
    raise SystemExit("NOTION_TOKEN is required")

NOTION_VERSION = "2022-06-28"
DB_MATERIALS = "43d801b4-a787-4101-bd12-d8b8199385c7"
DB_BOM = "6b67dc8d-bafb-49e8-9c39-bf66046a99fe"
DB_ORDERS = "50b7ce68-437e-431f-9a4f-a0d0d65a7b25"
DB_PICK_DETAIL = "267f16cf-e88a-41ed-8049-d39d618a1275"
DB_INBOUND = "cff100a4-ddcd-4bda-b8d7-57d44c4b3ce4"
DB_CORDERS = "64d6326e-c82a-4f5f-bccc-b34833f823c3"
DB_STOCK_LOG = "0aa78528-a5bb-4a0d-8005-c0c1e0aaf8a3"
DB_STOCK_EDIT_LOG = "808d789e-e9d3-4e6d-88d2-d58daa9aba4f"

MAT_STOCK = "C~uR"
MAT_SAFE = "JYP%3D"
MAT_CODE = "R%40aj"
MAT_UNIT = "f%7Dvw"
MAT_NOTE = "i%5BbD"
MAT_TYPE = "rWOy"
MAT_TITLE = "title"
BOM_QTY = "ElrP"
BOM_NOTE = "EokL"
BOM_CHILD = "uZv%3E"
BOM_PARENT = "yuq%3C"
BOM_TITLE = "title"

TYPE_FINISHED = "成品"
TYPE_SEMI = "半成品"
TYPE_PART = "零件"
UNIT_EACH = "個"
SYNC_NOTE = "BOM表修正.xlsx 同步 2026-07-17"

# (header cell, item column, marker column). The workbook intentionally uses
# several side-by-side BOM blocks rather than a normalized table.
BLOCKS = [
    ("A1", "B", "A"),
    ("D1", "E", "D"),
    ("G1", "H", "G"),
    ("L1", "M", "N"),
    ("O1", "O", "N"),
    ("R1", "R", "Q"),
    ("T1", "T", "V"),
    ("W1", "W", "V"),
    ("AA1", "AB", "AA"),
    ("AE1", "AE", "AD"),
    ("AH1", "AI", "AH"),
]

# Only explicit, reviewed aliases are merged. Broad zero-stripping is unsafe:
# e.g. Y-MW-11 and Y-MW-01-1 are different materials.
EXPLICIT_ALIASES = {
    "F-BRGK-02(LE)": ["F-BRGK-02"],
    "F-GAB-03D-B": ["F-GAB-3D-B"],
    "Y-SKC-01D-2": ["Y-SKC-1D-2", "SKC-01D-2", "SKC-1D-2"],
    "Y-SKC-23-1A": ["Y-SKC-23"],
    "Y-L23": ["Y-L-23"],
    "Y-FLT-D-01A": ["Y-FLT-D-1A"],
    "Y-FLT-D-01-1": ["Y-FLT-D-1-1"],
    "Y-FLT-D-01-1-1": ["Y-FLT-D-01-11"],
    "Y-SPA-09E": ["Y-SPA-9E"],
    "Y-SKC-07": ["Y-SKC-7"],
    "Y-SKC-04-1": ["Y-SKC-4-1"],
    "Y-SKC-02C#09": ["Y-SKC-2C#09", "SKC-02C#09", "SKC-2C#09"],
    "Y-MW-01-1B": ["Y-MW-1-1B", "MW-01-1B", "MW-1-1B"],
    "Y-MW-01-1A": ["Y-MW-1-1A", "MW-01-1A", "MW-1-1A"],
    "Y-MW-02": ["Y-MW-2", "MW-02", "MW-2"],
    "Y-MW-03": ["Y-MW-3", "MW-03", "MW-3"],
    "Y-MW-03-1": ["Y-MW-3-1", "MW-03-1", "MW-3-1"],
    "Y-MW-09": ["Y-MW-9", "MW-09", "MW-9"],
    "Z-FLT-C01": ["Z-FLT-C1", "FLT-C01", "FLT-C1"],
    "Z-FLT-B-C01": ["Z-FLT-B-C1", "FLT-B-C01", "FLT-B-C1"],
    "Z-FLT-B-C05": ["Z-FLT-B-C5", "FLT-B-C05", "FLT-B-C5"],
    "Z-FLT-D-C01": ["Z-FLT-D-C1", "Z-FLT-D-01", "FLT-D-C01"],
    "Z-FLT-B-R01-C1": ["Z-FLT-BR1-C1"],
    "Z-FLT-B-R-01A-C1": ["Z-FLT-BR1A-C1"],
    "Z-FLT-B-R02-C1": ["Z-FLT-BR2-C1"],
    "Z-FLT-BR02A-C1": ["Z-FLT-BR2A-C1", "Z-FLT-B-R02A-C1"],
}

# Reviewed against live stock logs on 2026-07-17. These values are not sums:
# the lower competing values belong to stale duplicate pages with no stock log.
STOCK_RESOLUTIONS = {
    "Y-SKC-04-1": 13544,
    "Y-SKC-07": 7504,
    "Z-FLT-D-C01": 2,
}


def normalize(value):
    return str(value or "").strip().upper()


def split_finished_label(value):
    text = str(value or "").strip()
    match = re.fullmatch(r"\s*([^()]+?)\s*\(([^()]*)\)\s*", text)
    if not match:
        return text, ""
    return match.group(1).strip(), match.group(2).strip()


def split_child_quantity(value):
    text = str(value or "").strip()
    match = re.fullmatch(r"(.+?)\*(\d+(?:\.\d+)?)", text)
    if not match:
        return text, 1
    qty = float(match.group(2))
    return match.group(1).strip(), int(qty) if qty.is_integer() else qty


def parse_workbook():
    if not SOURCE_XLSX.exists():
        raise SystemExit(f"Source workbook not found: {SOURCE_XLSX}")
    workbook = openpyxl.load_workbook(SOURCE_XLSX, data_only=True, read_only=False)
    sheet = workbook[workbook.sheetnames[0]]
    target = defaultdict(Counter)
    finished_codes = set()
    aliases = {key: list(values) for key, values in EXPLICIT_ALIASES.items()}

    for header_cell, item_col, marker_col in BLOCKS:
        canonical, old = split_finished_label(sheet[header_cell].value)
        if not canonical:
            continue
        finished_codes.add(canonical)
        if old and normalize(old) != normalize(canonical):
            aliases.setdefault(canonical, []).append(old)

        current_parent = None
        mode = None
        for row in range(2, sheet.max_row + 1):
            marker = str(sheet[f"{marker_col}{row}"].value or "").strip().replace(":", "")
            item = str(sheet[f"{item_col}{row}"].value or "").strip()
            if marker.startswith("成品"):
                current_parent = canonical
                mode = "finished"
                if item:
                    child, qty = split_child_quantity(item)
                    target[current_parent][child] += qty
                continue
            if marker.startswith("半成品"):
                current_parent = item or None
                mode = "semi"
                if current_parent:
                    target.setdefault(current_parent, Counter())
                continue
            if item and current_parent and mode:
                child, qty = split_child_quantity(item)
                target[current_parent][child] += qty

        target.setdefault(canonical, Counter())

    return {
        "target_bom": {parent: dict(children) for parent, children in target.items()},
        "finished_codes": finished_codes,
        "aliases": {code: sorted(set(values)) for code, values in aliases.items()},
    }


def api(method, endpoint, body=None, retries=8):
    url = "https://api.notion.com/v1/" + endpoint.lstrip("/")
    headers = {
        "Authorization": "Bearer " + TOKEN,
        "Notion-Version": NOTION_VERSION,
        "Content-Type": "application/json",
    }
    payload = None if body is None else json.dumps(body, ensure_ascii=False).encode("utf-8")
    for attempt in range(retries):
        req = request.Request(url, data=payload, headers=headers, method=method)
        try:
            with request.urlopen(req, timeout=90) as response:
                raw = response.read().decode("utf-8")
                return json.loads(raw) if raw else {}
        except error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            if exc.code == 429 or exc.code >= 500:
                retry_after = float(exc.headers.get("Retry-After") or 0)
                time.sleep(max(retry_after, 1.5 * (attempt + 1)))
                continue
            raise RuntimeError(f"{method} {endpoint} HTTP {exc.code}: {detail}") from exc
        except Exception:
            if attempt == retries - 1:
                raise
            time.sleep(1.2 * (attempt + 1))
    raise RuntimeError(f"{method} {endpoint} failed after retries")


def query_database(database_id, body=None):
    rows = []
    cursor = None
    while True:
        payload = {"page_size": 100, **(body or {})}
        if cursor:
            payload["start_cursor"] = cursor
        result = api("POST", f"databases/{database_id}/query", payload)
        rows.extend(result.get("results", []))
        if not result.get("has_more"):
            return rows
        cursor = result.get("next_cursor")
        time.sleep(0.36)


def property_by_id(properties, property_id):
    return next((prop for prop in properties.values() if prop.get("id") == property_id), {})


def title_text(prop):
    return "".join(item.get("plain_text", "") for item in prop.get("title", [])).strip()


def rich_text(prop):
    return "".join(item.get("plain_text", "") for item in prop.get("rich_text", [])).strip()


def select_name(prop):
    return (prop.get("select") or {}).get("name") or ""


def relation_ids(page, property_name):
    return [item["id"] for item in page.get("properties", {}).get(property_name, {}).get("relation", [])]


def material_from_page(page):
    props = page.get("properties", {})
    return {
        "id": page["id"],
        "url": page.get("url", ""),
        "name": title_text(property_by_id(props, MAT_TITLE)),
        "code": rich_text(property_by_id(props, MAT_CODE)),
        "type": select_name(property_by_id(props, MAT_TYPE)),
        "stock": property_by_id(props, MAT_STOCK).get("number") or 0,
        "safe": property_by_id(props, MAT_SAFE).get("number") or 0,
    }


def bom_from_page(page):
    props = page.get("properties", {})
    parents = property_by_id(props, BOM_PARENT).get("relation", [])
    children = property_by_id(props, BOM_CHILD).get("relation", [])
    return {
        "id": page["id"],
        "parent_id": parents[0]["id"] if parents else None,
        "child_id": children[0]["id"] if children else None,
        "qty": property_by_id(props, BOM_QTY).get("number"),
    }


def patch_page(page_id, properties=None, archived=None):
    body = {}
    if properties is not None:
        body["properties"] = properties
    if archived is not None:
        body["archived"] = archived
    result = api("PATCH", f"pages/{page_id}", body)
    time.sleep(0.36)
    return result


def expected_type(code, finished_codes, target_bom):
    if code in finished_codes or code.startswith("Z-"):
        return TYPE_FINISHED
    if code in target_bom or code.startswith("F-"):
        return TYPE_SEMI
    return TYPE_PART


def create_material(code, material_type):
    props = {
        MAT_TITLE: {"title": [{"text": {"content": code}}]},
        MAT_CODE: {"rich_text": [{"text": {"content": code}}]},
        MAT_TYPE: {"select": {"name": material_type}},
        MAT_UNIT: {"select": {"name": UNIT_EACH}},
        MAT_STOCK: {"number": 0},
        MAT_SAFE: {"number": 0},
        MAT_NOTE: {"rich_text": [{"text": {"content": SYNC_NOTE}}]},
    }
    return material_from_page(api("POST", "pages", {
        "parent": {"database_id": DB_MATERIALS},
        "properties": props,
    }))


def create_bom(parent, child, qty):
    parent_code = parent["code"] or parent["name"]
    child_code = child["code"] or child["name"]
    return api("POST", "pages", {
        "parent": {"database_id": DB_BOM},
        "properties": {
            BOM_TITLE: {"title": [{"text": {"content": f"{parent_code} -> {child_code}"}}]},
            BOM_PARENT: {"relation": [{"id": parent["id"]}]},
            BOM_CHILD: {"relation": [{"id": child["id"]}]},
            BOM_QTY: {"number": qty},
            BOM_NOTE: {"rich_text": [{"text": {"content": SYNC_NOTE}}]},
        },
    })


def candidate_materials(code, aliases, materials):
    values = {normalize(code), *(normalize(value) for value in aliases.get(code, []))}
    return [
        mat for mat in materials
        if normalize(mat["code"]) in values
        or (not normalize(mat["code"]) and normalize(mat["name"]) in values)
    ]


def usage_counts(material_ids, datasets):
    counts = Counter()
    ids = set(material_ids)
    for row in datasets["boms"]:
        if row["parent_id"] in ids:
            counts[row["parent_id"]] += 3
        if row["child_id"] in ids:
            counts[row["child_id"]] += 3
    for page, prop in datasets["orders"]:
        for item_id in relation_ids(page, prop):
            if item_id in ids:
                counts[item_id] += 2
    for key in ("picks", "inbounds"):
        for page, prop in datasets[key]:
            for item_id in relation_ids(page, prop):
                if item_id in ids:
                    counts[item_id] += 1
    for page in datasets["corders"]:
        value = rich_text(page.get("properties", {}).get("料件ID", {}))
        for item_id in ids:
            if item_id in value:
                counts[item_id] += 1
    return counts


def choose_keeper(code, candidates, usage):
    canonical = normalize(code)
    return sorted(candidates, key=lambda mat: (
        -usage.get(mat["id"], 0),
        0 if normalize(mat["code"]) == canonical else 1,
        0 if normalize(mat["name"]) == canonical else 1,
        0 if mat["stock"] else 1,
        mat["id"],
    ))[0]


def audit_stock_history(candidates):
    summaries = []
    for mat in candidates:
        code = mat["code"] or mat["name"]
        if not code:
            continue
        stock_rows = query_database(DB_STOCK_LOG, {
            "filter": {"property": "\u6599\u4ef6\u7de8\u865f", "rich_text": {"equals": code}},
            "sorts": [{"timestamp": "created_time", "direction": "descending"}],
        })
        edit_rows = query_database(DB_STOCK_EDIT_LOG, {
            "filter": {"property": "\u6599\u4ef6\u7de8\u865f", "rich_text": {"equals": code}},
            "sorts": [{"timestamp": "created_time", "direction": "descending"}],
        })
        summaries.append({
            "material_id": mat["id"],
            "code": code,
            "stock_events": [
                {
                    "created_time": row.get("created_time"),
                    "type": select_name(row.get("properties", {}).get("\u7570\u52d5\u985e\u578b", {})),
                    "qty": row.get("properties", {}).get("\u7570\u52d5\u6578\u91cf", {}).get("number"),
                    "before": row.get("properties", {}).get("\u7570\u52d5\u524d\u5eab\u5b58", {}).get("number"),
                    "after": row.get("properties", {}).get("\u7570\u52d5\u5f8c\u5eab\u5b58", {}).get("number"),
                    "ref": rich_text(row.get("properties", {}).get("\u95dc\u806f\u55ae\u865f", {})),
                }
                for row in stock_rows[:20]
            ],
            "edit_events": [
                {
                    "created_time": row.get("created_time"),
                    "type": select_name(row.get("properties", {}).get("\u4fee\u6539\u985e\u578b", {})),
                    "before": rich_text(row.get("properties", {}).get("\u4fee\u6539\u524d\u6578\u503c", {})),
                    "after": rich_text(row.get("properties", {}).get("\u4fee\u6539\u5f8c\u6578\u503c", {})),
                    "reason": rich_text(row.get("properties", {}).get("\u4fee\u6539\u539f\u56e0", {})),
                }
                for row in edit_rows[:20]
            ],
        })
    return summaries


def load_reference_datasets():
    bom_pages = query_database(DB_BOM)
    return {
        "bom_pages": bom_pages,
        "boms": [bom_from_page(page) for page in bom_pages],
        "orders": [(page, "成品") for page in query_database(DB_ORDERS)],
        "picks": [(page, "料件") for page in query_database(DB_PICK_DETAIL)],
        "inbounds": [(page, "料件") for page in query_database(DB_INBOUND)],
        "corders": query_database(DB_CORDERS),
    }


def replace_relation(page, property_name, replacements):
    current = relation_ids(page, property_name)
    next_ids = []
    for item_id in current:
        next_id = replacements.get(item_id, item_id)
        if next_id not in next_ids:
            next_ids.append(next_id)
    if next_ids == current:
        return False
    patch_page(page["id"], {property_name: {"relation": [{"id": item_id} for item_id in next_ids]}})
    return True


def migrate_references(replacements, datasets, selected_by_id):
    changed = Counter()
    for page in datasets["bom_pages"]:
        props = {}
        parents = relation_ids(page, "母件")
        children = relation_ids(page, "子件")
        next_parents = list(dict.fromkeys(replacements.get(item, item) for item in parents))
        next_children = list(dict.fromkeys(replacements.get(item, item) for item in children))
        if next_parents != parents:
            props["母件"] = {"relation": [{"id": item} for item in next_parents]}
        if next_children != children:
            props["子件"] = {"relation": [{"id": item} for item in next_children]}
        if props:
            patch_page(page["id"], props)
            changed["bom"] += 1

    for key in ("orders", "inbounds"):
        for page, prop in datasets[key]:
            if replace_relation(page, prop, replacements):
                changed[key] += 1

    for page, prop in datasets["picks"]:
        current = relation_ids(page, prop)
        next_ids = list(dict.fromkeys(replacements.get(item, item) for item in current))
        if next_ids == current:
            continue
        props = {prop: {"relation": [{"id": item} for item in next_ids]}}
        if next_ids:
            keep = selected_by_id.get(next_ids[0])
            if keep:
                props["料件名稱"] = {"title": [{"text": {"content": keep["code"] or keep["name"]}}]}
        patch_page(page["id"], props)
        changed["picks"] += 1

    for page in datasets["corders"]:
        prop = page.get("properties", {}).get("料件ID", {})
        current = rich_text(prop)
        updated = current
        for old_id, new_id in replacements.items():
            updated = updated.replace(old_id, new_id)
        if updated != current:
            patch_page(page["id"], {"料件ID": {"rich_text": [{"text": {"content": updated}}]}})
            changed["corders"] += 1
    return dict(changed)


def source_metadata():
    stat = SOURCE_XLSX.stat()
    return {
        "path": str(SOURCE_XLSX),
        "sheet": openpyxl.load_workbook(SOURCE_XLSX, read_only=True).sheetnames[0],
        "sha256": hashlib.sha256(SOURCE_XLSX.read_bytes()).hexdigest(),
        "size": stat.st_size,
        "modified": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(stat.st_mtime)),
    }


def write_report(report):
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")


def verify_no_references(old_ids):
    datasets = load_reference_datasets()
    found = []
    old = set(old_ids)
    for row in datasets["boms"]:
        if row["parent_id"] in old or row["child_id"] in old:
            found.append({"db": "bom", "page": row["id"]})
    for key in ("orders", "picks", "inbounds"):
        for page, prop in datasets[key]:
            if old.intersection(relation_ids(page, prop)):
                found.append({"db": key, "page": page["id"]})
    for page in datasets["corders"]:
        value = rich_text(page.get("properties", {}).get("料件ID", {}))
        if any(item_id in value for item_id in old):
            found.append({"db": "corders", "page": page["id"]})
    return found, datasets


def main():
    parser = argparse.ArgumentParser(description="Safely sync BOM表修正.xlsx to Notion")
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    parsed = parse_workbook()
    target_bom = parsed["target_bom"]
    finished_codes = parsed["finished_codes"]
    aliases = parsed["aliases"]
    all_codes = set(target_bom)
    for children in target_bom.values():
        all_codes.update(children)

    material_pages = query_database(DB_MATERIALS)
    materials = [material_from_page(page) for page in material_pages]
    datasets = load_reference_datasets()
    selected = {}
    plans = []
    blockers = []
    replacements = {}
    archive_ids = []

    for code in sorted(all_codes):
        candidates = candidate_materials(code, aliases, materials)
        if not candidates:
            plans.append({"action": "create_material", "code": code})
            continue
        usage = usage_counts([mat["id"] for mat in candidates], datasets)
        keeper = choose_keeper(code, candidates, usage)
        stocks = sorted(set(mat["stock"] for mat in candidates if mat["stock"] != 0))
        resolved_stock = STOCK_RESOLUTIONS.get(code)
        if len(stocks) > 1 and resolved_stock is None:
            blockers.append({
                "code": code,
                "reason": "conflicting_nonzero_stock",
                "candidates": candidates,
                "usage": {mat["id"]: usage.get(mat["id"], 0) for mat in candidates},
                "stock_history": audit_stock_history(candidates),
            })
            continue
        if resolved_stock is not None and resolved_stock not in stocks:
            blockers.append({
                "code": code,
                "reason": "reviewed_stock_resolution_no_longer_matches_live_data",
                "expected_stock": resolved_stock,
                "live_nonzero_stocks": stocks,
                "candidates": candidates,
            })
            continue
        target_stock = resolved_stock if resolved_stock is not None else (stocks[0] if stocks else keeper["stock"])
        remove = [mat for mat in candidates if mat["id"] != keeper["id"]]
        plans.append({
            "action": "normalize_material" if not remove else "merge_material",
            "code": code,
            "keeper": keeper,
            "remove": remove,
            "target_stock": target_stock,
            "stock_resolution": "reviewed_stock_log" if resolved_stock is not None else "single_nonzero_or_keeper",
            "usage": dict(usage),
        })
        for mat in remove:
            replacements[mat["id"]] = keeper["id"]
            archive_ids.append(mat["id"])
        selected[code] = keeper

    report = {
        "mode": "apply" if args.apply else "dry-run",
        "source": source_metadata(),
        "target": {
            "finished": sorted(finished_codes),
            "parents": len(target_bom),
            "materials": len(all_codes),
            "relations": sum(len(children) for children in target_bom.values()),
            "graph": target_bom,
        },
        "material_plans": plans,
        "blockers": blockers,
        "applied": {},
        "verification": {},
    }
    if blockers:
        write_report(report)
        print(json.dumps({"mode": report["mode"], "blockers": len(blockers), "report": str(REPORT_PATH)}, ensure_ascii=False))
        sys.exit(2)

    if not args.apply:
        action_counts = Counter(item["action"] for item in plans)
        report["planned_replacements"] = replacements
        write_report(report)
        print(json.dumps({
            "mode": "dry-run",
            "finished": len(finished_codes),
            "parents": len(target_bom),
            "materials": len(all_codes),
            "relations": sum(len(children) for children in target_bom.values()),
            "actions": dict(action_counts),
            "replacements": len(replacements),
            "report": str(REPORT_PATH),
        }, ensure_ascii=False))
        return

    # Ensure missing materials first, then normalize keeper pages in place so
    # most existing relations remain valid without any migration.
    for item in plans:
        if item["action"] == "create_material":
            code = item["code"]
            mat = create_material(code, expected_type(code, finished_codes, target_bom))
            materials.append(mat)
            selected[code] = mat
            time.sleep(0.36)
            continue
        code = item["code"]
        keeper = item["keeper"]
        properties = {}
        if keeper["code"] != code:
            properties[MAT_CODE] = {"rich_text": [{"text": {"content": code}}]}
        if keeper["name"] != code:
            properties[MAT_TITLE] = {"title": [{"text": {"content": code}}]}
        desired_type = expected_type(code, finished_codes, target_bom)
        if keeper["type"] != desired_type:
            properties[MAT_TYPE] = {"select": {"name": desired_type}}
        if keeper["stock"] != item["target_stock"]:
            properties[MAT_STOCK] = {"number": item["target_stock"]}
        if properties:
            properties[MAT_NOTE] = {"rich_text": [{"text": {"content": SYNC_NOTE}}]}
            patch_page(keeper["id"], properties)
        keeper.update({"code": code, "name": code, "type": desired_type, "stock": item["target_stock"]})
        selected[code] = keeper

    selected_by_id = {mat["id"]: mat for mat in selected.values()}
    migrated = migrate_references(replacements, datasets, selected_by_id) if replacements else {}

    # Re-read BOM after relation migration, then enforce the workbook's exact
    # direct-child graph for every parent represented in this correction file.
    bom_rows = [bom_from_page(page) for page in query_database(DB_BOM)]
    by_pair = defaultdict(list)
    for row in bom_rows:
        if row["parent_id"] and row["child_id"]:
            by_pair[(row["parent_id"], row["child_id"])].append(row)

    target_pairs = {}
    bom_actions = Counter()
    for parent_code, children in target_bom.items():
        parent = selected[parent_code]
        for child_code, qty in children.items():
            child = selected[child_code]
            pair = (parent["id"], child["id"])
            target_pairs[pair] = qty
            rows = by_pair.get(pair, [])
            if not rows:
                create_bom(parent, child, qty)
                time.sleep(0.36)
                bom_actions["created"] += 1
            else:
                keeper_row = rows[0]
                if keeper_row["qty"] != qty:
                    patch_page(keeper_row["id"], {
                        BOM_QTY: {"number": qty},
                        BOM_NOTE: {"rich_text": [{"text": {"content": SYNC_NOTE}}]},
                    })
                    bom_actions["updated"] += 1
                for duplicate in rows[1:]:
                    patch_page(duplicate["id"], archived=True)
                    bom_actions["archived_duplicate"] += 1

    target_parent_ids = {selected[code]["id"] for code in target_bom}
    for row in bom_rows:
        if row["parent_id"] in target_parent_ids and (row["parent_id"], row["child_id"]) not in target_pairs:
            patch_page(row["id"], archived=True)
            bom_actions["archived_stale"] += 1

    # Only archive aliases after every live relation has been migrated.
    dangling, _ = verify_no_references(archive_ids)
    if dangling:
        report["verification"] = {"passed": False, "dangling_before_archive": dangling}
        write_report(report)
        raise RuntimeError(f"Dangling references remain before archive: {len(dangling)}")
    for page_id in archive_ids:
        patch_page(page_id, archived=True)

    verify_materials = [material_from_page(page) for page in query_database(DB_MATERIALS)]
    verify_by_code = defaultdict(list)
    for mat in verify_materials:
        verify_by_code[normalize(mat["code"])].append(mat)
    verify_boms = [bom_from_page(page) for page in query_database(DB_BOM)]
    verify_pairs = defaultdict(list)
    for row in verify_boms:
        if row["parent_id"] and row["child_id"]:
            verify_pairs[(row["parent_id"], row["child_id"])].append(row)

    errors = []
    for code in sorted(all_codes):
        rows = verify_by_code[normalize(code)]
        if len(rows) != 1:
            errors.append({"type": "material_count", "code": code, "count": len(rows)})
    verified_selected = {code: verify_by_code[normalize(code)][0] for code in all_codes if len(verify_by_code[normalize(code)]) == 1}
    for parent_code, children in target_bom.items():
        parent = verified_selected.get(parent_code)
        if not parent:
            continue
        expected_ids = set()
        for child_code, qty in children.items():
            child = verified_selected.get(child_code)
            if not child:
                continue
            expected_ids.add(child["id"])
            rows = verify_pairs.get((parent["id"], child["id"]), [])
            if len(rows) != 1 or rows[0]["qty"] != qty:
                errors.append({
                    "type": "bom_pair",
                    "parent": parent_code,
                    "child": child_code,
                    "expected_qty": qty,
                    "rows": rows,
                })
        stale = [row for row in verify_boms if row["parent_id"] == parent["id"] and row["child_id"] not in expected_ids]
        if stale:
            errors.append({"type": "stale_bom", "parent": parent_code, "rows": stale})

    dangling_after, _ = verify_no_references(archive_ids)
    if dangling_after:
        errors.append({"type": "dangling_after_archive", "rows": dangling_after})

    report["applied"] = {
        "migrated_references": migrated,
        "archived_materials": len(archive_ids),
        "bom_actions": dict(bom_actions),
    }
    report["verification"] = {"passed": not errors, "errors": errors}
    write_report(report)
    print(json.dumps({
        "mode": "apply",
        "archived_materials": len(archive_ids),
        "migrated_references": migrated,
        "bom_actions": dict(bom_actions),
        "verification_passed": not errors,
        "errors": len(errors),
        "report": str(REPORT_PATH),
    }, ensure_ascii=False))
    if errors:
        sys.exit(2)


if __name__ == "__main__":
    main()
