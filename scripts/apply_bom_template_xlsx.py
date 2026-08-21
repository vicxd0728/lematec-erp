import argparse
import json
import os
import re
import time
import urllib.error
import urllib.request
from pathlib import Path

import openpyxl


WORKER = "https://green-wave-c22f.vic-e93.workers.dev"
NOTION_VERSION = "2022-06-28"
DB_MATERIALS = "43d801b4-a787-4101-bd12-d8b8199385c7"
DB_BOM = "6b67dc8d-bafb-49e8-9c39-bf66046a99fe"


def clean(v):
    return str(v or "").strip()


def sku(v):
    return re.sub(r"\s+", "", clean(v)).upper()


def parse_qty_part(raw):
    raw = clean(raw)
    m = re.match(r"^(.+?)(?:\s*(?:\*|x|X|×)\s*([0-9]+(?:\.[0-9]+)?))$", raw)
    if m:
        return sku(m.group(1)), float(m.group(2))
    return sku(raw), 1.0


def material_type(raw, code):
    raw = clean(raw).replace(" ", "")
    if raw in {"零件", "半成品", "成品", "蝦皮用", "蝦皮用(S-)"}:
        return raw
    if sku(code).startswith("S-"):
        return "蝦皮用"
    if "半" in raw:
        return "半成品"
    if "成" in raw:
        return "成品"
    return "零件"


def parse_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [clean(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows, errors = [], []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))}
        enabled = clean(data.get("是否啟用") or data.get("啟用") or data.get("匯入") or data.get("使用"))
        if enabled and enabled.lower() in {"否", "no", "n", "false", "0", "停用"}:
            continue
        parent = sku(data.get("母件料號") or data.get("母件") or data.get("成品") or data.get("料號"))
        bom_text = clean(data.get("BOM組成") or data.get("子件組成") or data.get("組成") or data.get("子件") or data.get("BOM"))
        note = clean(data.get("名稱/備註") or data.get("備註") or data.get("名稱"))
        typ = material_type(data.get("類型") or data.get("母件類型") or "", parent)
        if not parent and not bom_text:
            continue
        if not parent:
            errors.append(f"第 {idx} 行缺少母件料號")
            continue
        parts = []
        for raw in re.split(r"[\n,，、;；]+", bom_text):
            raw = clean(raw)
            if not raw:
                continue
            child, qty = parse_qty_part(raw)
            if child:
                parts.append({"child_sku": child, "quantity": qty, "raw": raw})
        if not parts:
            errors.append(f"第 {idx} 行 {parent} 沒有 BOM 組成")
            continue
        rows.append({"line": idx, "parent_sku": parent, "type": typ, "note": note, "parts": parts})
    return rows, errors


def http_json(method, url, body=None, token=None):
    data = None if body is None else json.dumps(body, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(url, data=data, method=method)
    req.add_header("Content-Type", "application/json")
    req.add_header("User-Agent", "Mozilla/5.0 LEMATEC-ERP-BOM-Template-Importer/1.0")
    if token:
        req.add_header("Authorization", f"Bearer {token}")
    for attempt in range(4):
        try:
            with urllib.request.urlopen(req, timeout=90) as resp:
                text = resp.read().decode("utf-8")
                return json.loads(text) if text else {}
        except urllib.error.HTTPError as exc:
            raw = exc.read().decode("utf-8", errors="replace")
            if exc.code == 429 and attempt < 3:
                time.sleep(1.5 * (attempt + 1))
                continue
            raise RuntimeError(f"{method} {url} HTTP {exc.code}: {raw[:500]}")


def notion(method, endpoint, token, body=None):
    url = "https://api.notion.com/v1/" + endpoint.lstrip("/")
    data = None if body is None else json.dumps(body, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(url, data=data, method=method)
    req.add_header("Authorization", f"Bearer {token}")
    req.add_header("Notion-Version", NOTION_VERSION)
    req.add_header("Content-Type", "application/json")
    for attempt in range(4):
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                return json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as exc:
            raw = exc.read().decode("utf-8", errors="replace")
            if exc.code == 429 and attempt < 3:
                time.sleep(1.5 * (attempt + 1))
                continue
            raise RuntimeError(f"Notion {method} {endpoint} HTTP {exc.code}: {raw[:500]}")


def archive_notion_page(page_id, token):
    if not page_id:
        return
    notion("PATCH", f"pages/{page_id}", token, {"archived": True})


def notion_all(db_id, token):
    out, cursor = [], None
    while True:
        body = {"page_size": 100}
        if cursor:
            body["start_cursor"] = cursor
        data = notion("POST", f"databases/{db_id}/query", token, body)
        out.extend(data.get("results") or [])
        if not data.get("has_more"):
            return out
        cursor = data.get("next_cursor")


def title(props, name):
    return "".join(x.get("plain_text", "") for x in ((props.get(name) or {}).get("title") or [])).strip()


def rich(props, name):
    return "".join(x.get("plain_text", "") for x in ((props.get(name) or {}).get("rich_text") or [])).strip()


def select(props, name):
    return (((props.get(name) or {}).get("select") or {}).get("name") or "").strip()


def relation_id(props, name):
    rel = (props.get(name) or {}).get("relation") or []
    return rel[0].get("id", "") if rel else ""


def number(props, name):
    val = (props.get(name) or {}).get("number")
    return float(val) if val is not None else 0.0


def load_notion_maps(token):
    mats = {}
    for p in notion_all(DB_MATERIALS, token):
        props = p.get("properties") or {}
        code = sku(rich(props, "料件編號") or title(props, "料件名稱"))
        if code and code not in mats:
            mats[code] = {"id": p["id"], "type": select(props, "類型")}
    bom = {}
    for p in notion_all(DB_BOM, token):
        props = p.get("properties") or {}
        parent, child = relation_id(props, "母件"), relation_id(props, "子件")
        if parent and child:
            bom[(parent, child)] = {"id": p["id"], "qty": number(props, "每台用量")}
    return mats, bom


def create_or_update_material(code, typ, note, token, existing):
    cur = existing.get(code)
    props = {
        "料件名稱": {"title": [{"text": {"content": code}}]},
        "料件編號": {"rich_text": [{"text": {"content": code}}]},
        "類型": {"select": {"name": typ or "零件"}},
        "單位": {"select": {"name": "個"}},
        "安全庫存": {"number": 0},
        "備註": {"rich_text": [{"text": {"content": note or ""}}]},
    }
    if cur:
        patch = {}
        if typ and cur.get("type") != typ:
            patch["類型"] = props["類型"]
        if note:
            patch["備註"] = props["備註"]
        if patch:
            notion("PATCH", f"pages/{cur['id']}", token, {"properties": patch})
        return cur["id"], False
    props["目前庫存"] = {"number": 0}
    page = notion("POST", "pages", token, {"parent": {"database_id": DB_MATERIALS}, "properties": props})
    existing[code] = {"id": page["id"], "type": typ or "零件"}
    return page["id"], True


def create_or_update_bom(parent_id, child_id, parent_sku, child_sku, qty, token, existing):
    cur = existing.get((parent_id, child_id))
    if cur:
        if abs(float(cur["qty"] or 0) - float(qty)) > 1e-9:
            notion("PATCH", f"pages/{cur['id']}", token, {"properties": {"每台用量": {"number": qty}, "備註": {"rich_text": [{"text": {"content": "BOM Excel qty sync"}}]}}})
            return cur["id"], "updated"
        return cur["id"], "skipped"
    title_text = f"{parent_sku} -> {child_sku}"
    page = notion(
        "POST",
        "pages",
        token,
        {
            "parent": {"database_id": DB_BOM},
            "properties": {
                "BOM項目": {"title": [{"text": {"content": title_text}}]},
                "母件": {"relation": [{"id": parent_id}]},
                "子件": {"relation": [{"id": child_id}]},
                "每台用量": {"number": qty},
                "備註": {"rich_text": [{"text": {"content": "BOM Excel import"}}]},
            },
        },
    )
    existing[(parent_id, child_id)] = {"id": page["id"], "qty": qty}
    return page["id"], "created"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--worker", default=WORKER)
    ap.add_argument("--replace-parent-boms", action="store_true")
    args = ap.parse_args()
    token = (os.environ.get("NOTION_TOKEN") or os.environ.get("NOTION_API_TOKEN") or "").strip()
    if not token:
        raise SystemExit("Missing NOTION_TOKEN or NOTION_API_TOKEN")

    parsed, parse_errors = parse_workbook(Path(args.xlsx))
    rows, pairs, errors, warnings = [], set(), [], []
    parent_type = {r["parent_sku"]: r["type"] for r in parsed}
    materials = {}
    for r in parsed:
        materials[r["parent_sku"]] = {"sku": r["parent_sku"], "name": r["parent_sku"], "material_type": r["type"], "stock": 0, "safety_stock": 0, "notes": r["note"]}
        for part in r["parts"]:
            child, qty = part["child_sku"], part["quantity"]
            pair = (r["parent_sku"], child)
            if r["parent_sku"] == child:
                warnings.append(f"第 {r['line']} 行移除自我關聯：{child}")
                continue
            typ = parent_type.get(child) or ("蝦皮用" if child.startswith("S-") else "零件")
            materials.setdefault(child, {"sku": child, "name": child, "material_type": typ, "stock": 0, "safety_stock": 0, "notes": f"BOM Excel子件，來源母件 {r['parent_sku']}"})
            if pair in pairs:
                for existing in rows:
                    if existing["parent_sku"] == r["parent_sku"] and existing["child_sku"] == child:
                        existing["quantity"] += qty
                        warnings.append(f"合併重複 BOM：{r['parent_sku']} -> {child}，累計用量 {existing['quantity']:g}")
                        break
                continue
            pairs.add(pair)
            rows.append({"parent_sku": r["parent_sku"], "child_sku": child, "quantity": qty, "notes": "BOM Excel import"})

    current_bom = http_json("GET", f"{args.worker}/api/inventory/bom/list?revision=preflight-{int(time.time())}")
    current_rows = current_bom.get("rows") or []
    parent_skus = {r["parent_sku"] for r in rows}
    current_pairs = {(sku(x.get("parent_sku")), sku(x.get("child_sku"))): x for x in current_rows}
    incoming_pairs = {(r["parent_sku"], r["child_sku"]) for r in rows}
    update_count = sum(1 for r in rows if (r["parent_sku"], r["child_sku"]) in current_pairs)
    create_count = len(rows) - update_count
    extra_rows = [x for (p, c), x in current_pairs.items() if p in parent_skus and (p, c) not in incoming_pairs]

    report = {
        "mode": "apply" if args.apply else "dry-run",
        "xlsx": str(args.xlsx),
        "parents": len(parsed),
        "bom_rows": len(rows),
        "materials_in_plan": len(materials),
        "new_pairs": create_count,
        "existing_pairs": update_count,
        "extra_existing_pairs_for_input_parents": len(extra_rows),
        "parse_errors": parse_errors,
        "errors": errors,
        "warnings": warnings,
        "extra_examples": extra_rows[:30],
    }
    if parse_errors or errors:
        print(json.dumps(report, ensure_ascii=False, indent=2))
        raise SystemExit(2)
    if not args.apply:
        print(json.dumps(report, ensure_ascii=False, indent=2))
        return

    payload = {"rows": rows, "materials": list(materials.values()), "replace_parent_boms": bool(args.replace_parent_boms)}
    supabase = http_json("POST", f"{args.worker}/api/inventory/bom/upsert", payload, token=token)

    notion_mats, notion_bom = load_notion_maps(token)
    material_page_ids, created_materials = {}, 0
    for code, spec in materials.items():
        pid, created = create_or_update_material(code, spec["material_type"], spec.get("notes", ""), token, notion_mats)
        material_page_ids[code] = pid
        created_materials += int(created)
        http_json("POST", f"{args.worker}/api/inventory/sync", {"kind": "upsert_material", "payload": {"sku": code, "name": code, "type": spec["material_type"], "stock": 0, "safe": 0, "note": spec.get("notes", ""), "notion_page_id": pid}})

    bom_created = bom_updated = bom_skipped = 0
    archived_extra_bom_mirrors = 0
    if args.replace_parent_boms:
        for extra in extra_rows:
            page_id = clean(extra.get("notion_page_id"))
            if page_id:
                archive_notion_page(page_id, token)
                archived_extra_bom_mirrors += 1
    rows_with_notion = []
    for r in rows:
        parent_id, child_id = material_page_ids[r["parent_sku"]], material_page_ids[r["child_sku"]]
        bom_page_id, status = create_or_update_bom(parent_id, child_id, r["parent_sku"], r["child_sku"], r["quantity"], token, notion_bom)
        if status == "created":
            bom_created += 1
        elif status == "updated":
            bom_updated += 1
        else:
            bom_skipped += 1
        rows_with_notion.append({**r, "notion_page_id": bom_page_id})

    http_json("POST", f"{args.worker}/api/inventory/bom/upsert", {"rows": rows_with_notion, "materials": list(materials.values()), "replace_parent_boms": bool(args.replace_parent_boms)}, token=token)
    verify = http_json("GET", f"{args.worker}/api/inventory/bom/list?revision=verify-{int(time.time())}")
    verify_pairs = {(sku(x.get("parent_sku")), sku(x.get("child_sku"))) for x in (verify.get("rows") or [])}
    missing_after = sorted([f"{p}->{c}" for p, c in incoming_pairs if (p, c) not in verify_pairs])
    report.update(
        {
            "supabase": supabase,
            "notion_created_materials": created_materials,
            "notion_archived_extra_bom_mirrors": archived_extra_bom_mirrors,
            "notion_bom_created": bom_created,
            "notion_bom_updated": bom_updated,
            "notion_bom_skipped": bom_skipped,
            "verify_missing_after": missing_after[:50],
            "verify_missing_count": len(missing_after),
        }
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
